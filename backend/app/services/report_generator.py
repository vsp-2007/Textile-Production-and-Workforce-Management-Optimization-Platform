"""
Report Generator Service
Generates various production reports and exports
"""

from datetime import datetime, date, timedelta, time
from sqlalchemy import func, and_, or_
from io import BytesIO

from app import db
from app.models import (
    DailyReport, Shift, ShiftAssignment, AssignmentStatus,
    ProductionLog, Machine, MachineDowntime, MachineStatus,
    User, UserRole, MachineType, Department
)


def generate_shift_report_data(shift_id, report_date):
    """Generate detailed report for a specific shift on a date"""
    shift = Shift.query.get(shift_id)
    if not shift:
        return None
    
    # Get all assignments for this shift
    assignments = ShiftAssignment.query.filter_by(shift_id=shift_id).all()
    
    # Get production logs for these assignments
    assignment_ids = [a.id for a in assignments]
    logs = ProductionLog.query.filter(ProductionLog.shift_assignment_id.in_(assignment_ids)).all() if assignment_ids else []
    
    # Get downtime for machines in this shift
    machine_ids = [a.machine_id for a in assignments]
    downtimes = MachineDowntime.query.filter(
        MachineDowntime.machine_id.in_(machine_ids),
        MachineDowntime.start_time >= datetime.combine(report_date, time.min),
        MachineDowntime.start_time < datetime.combine(report_date + timedelta(days=1), time.min)
    ).all() if machine_ids else []
    
    # Calculate metrics
    total_machines = len(machine_ids)
    active_machines = len([a for a in assignments if a.status in [AssignmentStatus.STARTED, AssignmentStatus.COMPLETED]])
    total_operators = len(set(a.operator_id for a in assignments))
    present_operators = len(set(a.operator_id for a in assignments if a.status in [AssignmentStatus.STARTED, AssignmentStatus.COMPLETED]))
    
    total_yards = sum(log.actual_yards or 0 for log in logs)
    total_waste = sum(log.waste_yards or 0 for log in logs)
    total_target = sum(log.target_yards or 0 for log in logs)
    
    downtime_minutes = sum(d.duration_minutes or 0 for d in downtimes)
    
    # OEE calculation per machine
    machine_oee = {}
    for machine_id in machine_ids:
        machine_logs = [l for l in logs if l.machine_id == machine_id]
        machine_downtimes = [d for d in downtimes if d.machine_id == machine_id]
        
        m_yards = sum(l.actual_yards or 0 for l in machine_logs)
        m_target = sum(l.target_yards or 0 for l in machine_logs)
        m_waste = sum(l.waste_yards or 0 for l in machine_logs)
        m_downtime = sum(d.duration_minutes or 0 for d in machine_downtimes)
        
        machine = Machine.query.get(machine_id)
        planned_minutes = shift.duration_hours * 60
        operating_minutes = planned_minutes - m_downtime
        
        availability = (operating_minutes / planned_minutes * 100) if planned_minutes > 0 else 0
        performance = (m_yards / m_target * 100) if m_target > 0 else 0
        good_yards = sum(l.actual_yards or 0 for l in machine_logs if l.quality_grade in ['A', 'B'])
        quality = (good_yards / m_yards * 100) if m_yards > 0 else 0
        oee = (availability * performance * quality) / 10000
        
        machine_oee[machine_id] = {
            'machine_code': machine.machine_code if machine else 'Unknown',
            'availability': round(availability, 1),
            'performance': round(performance, 1),
            'quality': round(quality, 1),
            'oee': round(oee, 1),
            'yards': m_yards,
            'target': m_target,
            'waste': m_waste,
            'downtime_minutes': m_downtime
        }
    
    avg_oee = sum(m['oee'] for m in machine_oee.values()) / len(machine_oee) if machine_oee else 0
    
    return {
        'shift': shift.to_dict(),
        'report_date': report_date.isoformat(),
        'summary': {
            'total_machines': total_machines,
            'active_machines': active_machines,
            'total_operators': total_operators,
            'present_operators': present_operators,
            'total_yards': total_yards,
            'total_waste': total_waste,
            'total_target': total_target,
            'efficiency': round(total_yards / total_target * 100, 1) if total_target > 0 else 0,
            'waste_percentage': round(total_waste / total_yards * 100, 1) if total_yards > 0 else 0,
            'avg_oee': round(avg_oee, 1),
            'downtime_minutes': downtime_minutes
        },
        'assignments': [a.to_dict() for a in assignments],
        'production_logs': [l.to_dict() for l in logs],
        'downtimes': [d.to_dict() for d in downtimes],
        'machine_oee': machine_oee,
        'generated_at': datetime.utcnow().isoformat()
    }


def generate_daily_report_data(report_date):
    """Generate daily report for all shifts"""
    shifts = Shift.query.filter_by(is_active=True).all()
    
    reports = []
    total_yards = 0
    total_waste = 0
    total_target = 0
    total_downtime = 0
    all_oee = []
    
    for shift in shifts:
        shift_report = generate_shift_report_data(shift.id, report_date)
        if shift_report:
            reports.append(shift_report)
            total_yards += shift_report['summary']['total_yards']
            total_waste += shift_report['summary']['total_waste']
            total_target += shift_report['summary']['total_target']
            total_downtime += shift_report['summary']['downtime_minutes']
            all_oee.append(shift_report['summary']['avg_oee'])
    
    avg_oee = sum(all_oee) / len(all_oee) if all_oee else 0
    
    return {
        'report_date': report_date.isoformat(),
        'summary': {
            'total_yards': total_yards,
            'total_waste': total_waste,
            'total_target': total_target,
            'overall_efficiency': round(total_yards / total_target * 100, 1) if total_target > 0 else 0,
            'waste_percentage': round(total_waste / total_yards * 100, 1) if total_yards > 0 else 0,
            'avg_oee': round(avg_oee, 1),
            'total_downtime_minutes': total_downtime,
            'total_shifts': len(shifts)
        },
        'shifts': reports,
        'generated_at': datetime.utcnow().isoformat()
    }


def generate_machine_report_data(machine_id, days=30):
    """Generate machine performance report"""
    machine = Machine.query.get(machine_id)
    if not machine:
        return None
    
    since = datetime.utcnow() - timedelta(days=days)
    
    # Production logs
    logs = ProductionLog.query.filter(
        ProductionLog.machine_id == machine_id,
        ProductionLog.start_time >= since
    ).order_by(ProductionLog.start_time).all()
    
    # Downtimes
    downtimes = MachineDowntime.query.filter(
        MachineDowntime.machine_id == machine_id,
        MachineDowntime.start_time >= since
    ).order_by(MachineDowntime.start_time).all()
    
    # Telemetry
    from app.models import MachineTelemetry
    telemetry = MachineTelemetry.query.filter(
        MachineTelemetry.machine_id == machine_id,
        MachineTelemetry.timestamp >= since
    ).order_by(MachineTelemetry.timestamp.desc()).limit(1000).all()
    
    total_yards = sum(l.actual_yards or 0 for l in logs)
    total_waste = sum(l.waste_yards or 0 for l in logs)
    total_target = sum(l.target_yards or 0 for l in logs)
    total_downtime = sum(d.duration_minutes or 0 for d in downtimes)
    
    # Daily breakdown
    daily = {}
    for log in logs:
        day = log.start_time.date().isoformat()
        if day not in daily:
            daily[day] = {'yards': 0, 'waste': 0, 'target': 0, 'logs': 0}
        daily[day]['yards'] += log.actual_yards or 0
        daily[day]['waste'] += log.waste_yards or 0
        daily[day]['target'] += log.target_yards or 0
        daily[day]['logs'] += 1
    
    daily_list = [
        {
            'date': k,
            'yards': v['yards'],
            'waste': v['waste'],
            'target': v['target'],
            'efficiency': round(v['yards'] / v['target'] * 100, 1) if v['target'] > 0 else 0
        }
        for k, v in sorted(daily.items())
    ]
    
    # Status distribution from telemetry
    status_counts = {}
    for t in telemetry:
        status_counts[t.status.value] = status_counts.get(t.status.value, 0) + 1
    
    return {
        'machine': machine.to_dict(),
        'period_days': days,
        'summary': {
            'total_yards': total_yards,
            'total_waste': total_waste,
            'total_target': total_target,
            'efficiency': round(total_yards / total_target * 100, 1) if total_target > 0 else 0,
            'waste_percentage': round(total_waste / total_yards * 100, 1) if total_yards > 0 else 0,
            'total_downtime_minutes': total_downtime,
            'total_shifts': len(logs),
            'status_distribution': status_counts
        },
        'daily': daily_list,
        'downtimes': [d.to_dict() for d in downtimes],
        'recent_telemetry': [t.to_dict() for t in telemetry[:50]],
        'generated_at': datetime.utcnow().isoformat()
    }


def generate_operator_report_data(operator_id, days=30):
    """Generate operator productivity report"""
    operator = User.query.filter_by(id=operator_id, role=UserRole.OPERATOR).first()
    if not operator:
        return None
    
    since = datetime.utcnow() - timedelta(days=days)
    
    # Assignments
    assignments = ShiftAssignment.query.filter(
        ShiftAssignment.operator_id == operator_id,
        ShiftAssignment.assigned_at >= since
    ).order_by(ShiftAssignment.assigned_at).all()
    
    # Production logs
    assignment_ids = [a.id for a in assignments]
    logs = ProductionLog.query.filter(ProductionLog.shift_assignment_id.in_(assignment_ids)).all() if assignment_ids else []
    
    total_yards = sum(l.actual_yards or 0 for l in logs)
    total_waste = sum(l.waste_yards or 0 for l in logs)
    total_target = sum(l.target_yards or 0 for l in logs)
    
    completed_shifts = len([a for a in assignments if a.status == AssignmentStatus.COMPLETED])
    total_shifts = len(assignments)
    total_hours = sum(a.shift.duration_hours for a in assignments if a.shift)
    
    # By machine
    machine_stats = {}
    for log in logs:
        mid = log.machine_id
        if mid not in machine_stats:
            machine_stats[mid] = {'yards': 0, 'waste': 0, 'target': 0, 'shifts': 0}
        machine_stats[mid]['yards'] += log.actual_yards or 0
        machine_stats[mid]['waste'] += log.waste_yards or 0
        machine_stats[mid]['target'] += log.target_yards or 0
        machine_stats[mid]['shifts'] += 1
    
    machine_list = []
    for mid, stats in machine_stats.items():
        machine = Machine.query.get(mid)
        if machine:
            machine_list.append({
                'machine_id': mid,
                'machine_code': machine.machine_code,
                'machine_name': machine.name,
                **stats,
                'efficiency': round(stats['yards'] / stats['target'] * 100, 1) if stats['target'] > 0 else 0
            })
    
    # Certifications
    certs = OperatorCertification.query.filter_by(user_id=operator_id).all()
    
    return {
        'operator': operator.to_dict(),
        'period_days': days,
        'summary': {
            'total_shifts': total_shifts,
            'completed_shifts': completed_shifts,
            'total_hours': total_hours,
            'total_yards': total_yards,
            'total_waste': total_waste,
            'total_target': total_target,
            'efficiency': round(total_yards / total_target * 100, 1) if total_target > 0 else 0,
            'waste_percentage': round(total_waste / total_yards * 100, 1) if total_yards > 0 else 0,
            'yards_per_hour': round(total_yards / total_hours, 1) if total_hours > 0 else 0
        },
        'assignments': [a.to_dict() for a in assignments],
        'production_logs': [l.to_dict() for l in logs],
        'by_machine': machine_list,
        'certifications': [c.to_dict() for c in certs],
        'generated_at': datetime.utcnow().isoformat()
    }


def export_report_pdf(report_data, report_type):
    """Export report as PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSample
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.units import inch
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = f"{report_type.replace('_', ' ').title()} Report"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    
    # Summary table
    if 'summary' in report_data:
        summary = report_data['summary']
        summary_data = [['Metric', 'Value']]
        for key, value in summary.items():
            if isinstance(value, dict):
                continue
            summary_data.append([key.replace('_', ' ').title(), str(value)])
        
        table = Table(summary_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_report_excel(report_data, report_type):
    """Export report as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    
    # Headers
    headers = ['Metric', 'Value']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Summary data
    if 'summary' in report_data:
        row = 2
        for key, value in report_data['summary'].items():
            if isinstance(value, dict):
                continue
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=value)
            row += 1
    
    # Detail sheets based on report type
    if report_type == 'daily' and 'shifts' in report_data:
        for i, shift_report in enumerate(report_data['shifts']):
            ws_shift = wb.create_sheet(title=f"Shift_{shift_report['shift']['code'][:20]}")
            _write_shift_detail(ws_shift, shift_report)
    
    elif report_type == 'shift':
        _write_shift_detail(ws, report_data)
    
    elif report_type == 'machine':
        _write_machine_detail(wb, report_data)
    
    elif report_type == 'operator':
        _write_operator_detail(wb, report_data)
    
    # Auto-fit columns
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _write_shift_detail(ws, shift_report):
    """Write shift detail to worksheet"""
    from openpyxl.styles import Font, PatternFill
    
    # Assignment headers
    headers = ['Machine', 'Operator', 'Status', 'Target (yds)', 'Actual (yds)', 'Waste (yds)', 'Efficiency (%)', 'Quality']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Assignment data
    assignments = shift_report.get('assignments', [])
    logs_by_assignment = {}
    for log in shift_report.get('production_logs', []):
        logs_by_assignment[log['shift_assignment_id']] = log
    
    for row, assignment in enumerate(assignments, 2):
        log = logs_by_assignment.get(assignment['id'], {})
        ws.cell(row=row, column=1, value=assignment.get('machine', {}).get('machine_code', ''))
        ws.cell(row=row, column=2, value=assignment.get('operator', {}).get('name', ''))
        ws.cell(row=row, column=3, value=assignment.get('status', ''))
        ws.cell(row=row, column=4, value=log.get('target_yards', 0))
        ws.cell(row=row, column=5, value=log.get('actual_yards', 0))
        ws.cell(row=row, column=6, value=log.get('waste_yards', 0))
        ws.cell(row=row, column=7, value=log.get('efficiency', 0))
        ws.cell(row=row, column=8, value=log.get('quality_grade', ''))


def _write_machine_detail(wb, report_data):
    """Write machine detail to workbook"""
    from openpyxl.styles import Font, PatternFill
    
    # Daily production
    ws_daily = wb.create_sheet(title="Daily Production")
    headers = ['Date', 'Target (yds)', 'Actual (yds)', 'Waste (yds)', 'Efficiency (%)']
    for col, header in enumerate(headers, 1):
        cell = ws_daily.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for row, day in enumerate(report_data.get('daily', []), 2):
        ws_daily.cell(row=row, column=1, value=day['date'])
        ws_daily.cell(row=row, column=2, value=day['target'])
        ws_daily.cell(row=row, column=3, value=day['yards'])
        ws_daily.cell(row=row, column=4, value=day['waste'])
        ws_daily.cell(row=row, column=5, value=day['efficiency'])
    
    # Downtimes
    ws_downtime = wb.create_sheet(title="Downtimes")
    headers = ['Start', 'End', 'Reason', 'Duration (min)', 'Reported By', 'Resolved By']
    for col, header in enumerate(headers, 1):
        cell = ws_downtime.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for row, dt in enumerate(report_data.get('downtimes', []), 2):
        ws_downtime.cell(row=row, column=1, value=dt.get('start_time', ''))
        ws_downtime.cell(row=row, column=2, value=dt.get('end_time', ''))
        ws_downtime.cell(row=row, column=3, value=dt.get('reason', ''))
        ws_downtime.cell(row=row, column=4, value=dt.get('duration_minutes', 0))
        ws_downtime.cell(row=row, column=5, value=dt.get('reported_by', ''))
        ws_downtime.cell(row=row, column=6, value=dt.get('resolved_by', ''))


def _write_operator_detail(wb, report_data):
    """Write operator detail to workbook"""
    from openpyxl.styles import Font, PatternFill
    
    # Assignments
    ws_assign = wb.create_sheet(title="Assignments")
    headers = ['Date', 'Shift', 'Machine', 'Status', 'Target', 'Actual', 'Waste', 'Efficiency', 'Quality']
    for col, header in enumerate(headers, 1):
        cell = ws_assign.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    logs_by_assignment = {log['shift_assignment_id']: log for log in report_data.get('production_logs', [])}
    
    for row, assignment in enumerate(report_data.get('assignments', []), 2):
        log = logs_by_assignment.get(assignment['id'], {})
        ws_assign.cell(row=row, column=1, value=assignment.get('assigned_at', '')[:10])
        ws_assign.cell(row=row, column=2, value=assignment.get('shift', {}).get('name', ''))
        ws_assign.cell(row=row, column=3, value=assignment.get('machine', {}).get('machine_code', ''))
        ws_assign.cell(row=row, column=4, value=assignment.get('status', ''))
        ws_assign.cell(row=row, column=5, value=log.get('target_yards', 0))
        ws_assign.cell(row=row, column=6, value=log.get('actual_yards', 0))
        ws_assign.cell(row=row, column=7, value=log.get('waste_yards', 0))
        ws_assign.cell(row=row, column=8, value=log.get('efficiency', 0))
        ws_assign.cell(row=row, column=9, value=log.get('quality_grade', ''))
    
    # By machine
    ws_machine = wb.create_sheet(title="By Machine")
    headers = ['Machine', 'Shifts', 'Target (yds)', 'Actual (yds)', 'Waste (yds)', 'Efficiency (%)']
    for col, header in enumerate(headers, 1):
        cell = ws_machine.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for row, machine in enumerate(report_data.get('by_machine', []), 2):
        ws_machine.cell(row=row, column=1, value=f"{machine['machine_code']} - {machine['machine_name']}")
        ws_machine.cell(row=row, column=2, value=machine['shifts'])
        ws_machine.cell(row=row, column=3, value=machine['target'])
        ws_machine.cell(row=row, column=4, value=machine['yards'])
        ws_machine.cell(row=row, column=5, value=machine['waste'])
        ws_machine.cell(row=row, column=6, value=machine['efficiency'])


def generate_daily_report(shift_id, report_date):
    """Generate and save daily report to database"""
    if isinstance(report_date, str):
        report_date = datetime.strptime(report_date, '%Y-%m-%d').date()
    
    shift = Shift.query.get(shift_id)
    if not shift:
        return None
    
    report_data = generate_shift_report_data(shift_id, report_date)
    
    # Check if exists
    existing = DailyReport.query.filter_by(report_date=report_date, shift_id=shift_id).first()
    
    if existing:
        existing.total_machines = report_data['summary']['total_machines']
        existing.active_machines = report_data['summary']['active_machines']
        existing.total_operators = report_data['summary']['total_operators']
        existing.present_operators = report_data['summary']['present_operators']
        existing.total_yards = report_data['summary']['total_yards']
        existing.total_waste = report_data['summary']['total_waste']
        existing.avg_oee = report_data['summary']['avg_oee']
        existing.downtime_minutes = report_data['summary']['downtime_minutes']
        existing.generated_at = datetime.utcnow()
    else:
        existing = DailyReport(
            report_date=report_date,
            shift_id=shift_id,
            total_machines=report_data['summary']['total_machines'],
            active_machines=report_data['summary']['active_machines'],
            total_operators=report_data['summary']['total_operators'],
            present_operators=report_data['summary']['present_operators'],
            total_yards=report_data['summary']['total_yards'],
            total_waste=report_data['summary']['total_waste'],
            avg_oee=report_data['summary']['avg_oee'],
            downtime_minutes=report_data['summary']['downtime_minutes']
        )
        db.session.add(existing)
    
    db.session.commit()
    return existing